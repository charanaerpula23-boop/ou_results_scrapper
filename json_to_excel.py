#!/usr/bin/env python3
"""Convert result JSON/NDJSON to an Excel workbook with 'students' and 'marks' sheets.

Usage:
  python json_to_excel.py input.json output.xlsx

If `input.json` is NDJSON (one JSON object per line) it will be handled.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import pandas as pd
except Exception:
    pd = None


def load_items(path: Path):
    txt = path.read_text(encoding="utf-8").strip()
    if not txt:
        return []
    if txt.startswith("["):
        return json.loads(txt)
    items = []
    for line in txt.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except Exception:
            # skip malformed lines
            continue
    return items


def items_to_excel(items, out_path: Path, protect_password: str = None):
    if pd is None:
        print("Error: pandas and openpyxl are required. Install: python -m pip install pandas openpyxl")
        sys.exit(2)

    # Helper to flatten top-level dicts (skip lists)
    def flatten_top(item):
        flat = {}
        for k, v in item.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat[f"{k}.{k2}"] = v2
            elif isinstance(v, list):
                # skip lists (handled separately)
                continue
            else:
                flat[k] = v
        return flat

    rows = []
    marks = []
    for item in items:
        # Flatten top-level keys (student.* becomes columns, etc.)
        flat = flatten_top(item)
        # Add explicit top-level hallticket for joining later
        flat["hallticket"] = (item.get("student") or {}).get("hallticket")
        # Do NOT include raw JSON in results
        rows.append(flat)

        # Marks: collect all marks (if any)
        for m in item.get("marks", []):
            mark_row = {"hallticket": (item.get("student") or {}).get("hallticket")}
            for mk, mv in (m or {}).items():
                mark_row[mk] = mv
            marks.append(mark_row)

    df_rows = pd.DataFrame(rows)
    df_marks = pd.DataFrame(marks)

    # Ensure expected mark columns exist and are first/ordered
    expected_mark_cols = ["hallticket", "code", "subject", "credits", "grade"]
    if df_marks.empty:
        df_marks = pd.DataFrame(columns=expected_mark_cols)
    else:
        df_marks = df_marks.reindex(columns=expected_mark_cols).fillna("")

    # Pivot marks to create per-subject columns for grade, credits, subject (if any marks exist)
    if not df_marks.empty:
        # grade pivots keyed by code (codes as column names)
        df_grade_codes = df_marks.pivot_table(index="hallticket", columns="code", values="grade", aggfunc="first")
        df_credits_pivot = df_marks.pivot_table(index="hallticket", columns="code", values="credits", aggfunc="first")
        df_subject_pivot = df_marks.pivot_table(index="hallticket", columns="code", values="subject", aggfunc="first")

        # Merge pivots into df_rows on hallticket
        df_rows = df_rows.merge(df_grade_codes.reset_index(), on="hallticket", how="left")
        df_rows = df_rows.merge(df_credits_pivot.reset_index().rename(columns=lambda x: f"credits_{x}" if x != 'hallticket' else x), on="hallticket", how="left")
        df_rows = df_rows.merge(df_subject_pivot.reset_index().rename(columns=lambda x: f"subject_{x}" if x != 'hallticket' else x), on="hallticket", how="left")

    # Reorder columns: exactly -> hallticket, name, father, subject-code columns, then result
    core_pref = ["hallticket", "student.name", "student.father"]
    core_cols = [c for c in core_pref if c in df_rows.columns]

    # Determine codes ordered by frequency (descending)
    code_counts = df_marks['code'].value_counts() if not df_marks.empty else pd.Series([], dtype=int)
    codes_sorted = list(code_counts.index) if len(code_counts) else []

    # Separate lab codes which should appear between common and uncommon subjects
    labs = [c for c in codes_sorted if 'LAB' in c]
    non_labs = [c for c in codes_sorted if c not in labs]

    # Split non-labs into common and uncommon halves (common first)
    half = len(non_labs) // 2
    common = non_labs[:half]
    uncommon = non_labs[half:]

    # Final subject code order: common, labs, uncommon (only include codes present in df_rows)
    subject_codes_ordered = [c for c in (common + labs + uncommon) if c in df_rows.columns]

    # Ensure 'result' is appended at the end
    final_cols = core_cols + subject_codes_ordered
    if 'result' in df_rows.columns:
        final_cols.append('result')

    # Add any missing subject-code columns (alphabetically) that might exist (as fallback)
    remaining_code_cols = [c for c in df_rows.columns if c not in final_cols and isinstance(c, str) and c.isupper() and not c.startswith(('credits_','subject_'))]
    remaining_code_cols = sorted(remaining_code_cols)
    final_cols = final_cols + remaining_code_cols

    # Fill missing columns with empty strings and reindex
    for c in final_cols:
        if c not in df_rows.columns:
            df_rows[c] = ""
    df_rows = df_rows.reindex(columns=final_cols)
    df_rows = df_rows.fillna("")

    # Sort by hallticket ascending when possible
    try:
        df_rows['hallticket_sort'] = df_rows['hallticket'].astype(int)
        df_rows = df_rows.sort_values('hallticket_sort').drop(columns=['hallticket_sort'])
    except Exception:
        df_rows = df_rows.sort_values('hallticket')

    # Prepare raw sheet
    raw_df = pd.DataFrame({"json": [json.dumps(i, ensure_ascii=False) for i in items]})

    # Write sheets: results (final ordered), marks (long), grades (pivot), raw (original JSON lines)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_rows.to_excel(writer, sheet_name="results", index=False)
        if not df_marks.empty:
            # write a grades sheet that shows grades per code (codes as columns)
            df_grade_only = df_grade_codes.reset_index()
            df_grade_only.to_excel(writer, sheet_name="grades", index=False)
        else:
            pd.DataFrame(columns=["hallticket"]).to_excel(writer, sheet_name="grades", index=False)
        df_marks.to_excel(writer, sheet_name="marks", index=False)
        raw_df.to_excel(writer, sheet_name="raw", index=False)

    # Optionally protect the grades sheet (lock grade cells)
    if protect_password:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Protection
        except Exception:
            print("Error: openpyxl required for protecting the spreadsheet.")
            return

        wb = load_workbook(out_path)
        if "grades" in wb.sheetnames:
            ws = wb["grades"]
            # Unlock all cells first
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)
            # Lock grade cells (all except header and hallticket column)
            max_col = ws.max_column
            max_row = ws.max_row
            # find hallticket column index (assumed first column)
            for r in range(2, max_row + 1):
                for c in range(2, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.protection = Protection(locked=True)
            ws.protection.sheet = True
            ws.protection.set_password(protect_password)
            wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Convert JSON/NDJSON results to an Excel workbook.")
    parser.add_argument("input", help="Input JSON or NDJSON file")
    parser.add_argument("output", nargs="?", help="Output .xlsx file (defaults to input with .xlsx)")
    parser.add_argument("--protect-password", help="Password to protect the grades sheet (optional)", default=None)
    args = parser.parse_args()

    protect_password = args.protect_password

    inp = Path(args.input)
    if not inp.exists():
        print(f"Input file not found: {inp}")
        sys.exit(1)

    out = Path(args.output) if args.output else inp.with_suffix(".xlsx")

    items = load_items(inp)
    if not items:
        print("No data found in input file.")
        sys.exit(1)

    items_to_excel(items, out, protect_password)
    print(f"Done — Excel saved to {out}")


if __name__ == "__main__":
    main()
