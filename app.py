import requests
from bs4 import BeautifulSoup
import json
import os
import urllib3
import threading
import argparse
import sys
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# Excel export (required)
try:
    import pandas as pd
    import openpyxl  # ensure engine available
except Exception:
    # Fail fast with clear instruction
    sys.exit("Error: pandas and openpyxl are required. Install: python -m pip install -r requirements.txt")

# ================= SSL FIX =================
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= CONFIG =================
# Default configuration (can be overridden via CLI flags or interactive prompts)
URL = "https://www.osmania.ac.in/res07/20250686.jsp"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Referer": URL,
    "Content-Type": "application/x-www-form-urlencoded",
}

START_HT = 110624861001
END_HT = 110624861064

MAX_WORKERS = 15  # aggressive
OUTPUT_FILE = "ou_results.xlsx"
# Temporary NDJSON file used while incrementally appending results (final output will be Excel)
NDJSON_FILE = "ou_results.ndjson"

# Auto-excel options
AUTO_EXCEL = True
AUTO_EXCEL_INTERVAL = 10  # seconds
PROTECT_PASSWORD = None

# Internal sync primitives for auto-excel
auto_excel_event = threading.Event()
auto_excel_stop = threading.Event()
auto_excel_thread = None

# =========================================

session = requests.Session()
lock = threading.Lock()
results = []


def clear_old_results():
    # Truncate NDJSON staging file so we can append NDJSON lines as results arrive
    with open(NDJSON_FILE, "w", encoding="utf-8") as f:
        f.write("")


def append_result(data):
    # Append a single JSON object as one line (NDJSON) and fsync to disk
    with lock:
        with open(NDJSON_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(data, ensure_ascii=False) + "\n")
            f.flush()
            try:
                os.fsync(f.fileno())
            except Exception:
                pass
    # signal the auto-excel worker that new data is available
    if AUTO_EXCEL:
        auto_excel_event.set()


def save_results():
    # Optional: write the in-memory results as a JSON array (overwrites file)
    with lock:
        with open(NDJSON_FILE, "w", encoding="utf-8") as f:
            for item in results:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")


def generate_excel():
    """Generate the Excel workbook from the current NDJSON staging file.
    Uses the helper in json_to_excel.py for consistent output.
    """
    try:
        # import lazily to avoid circular or heavy imports at module load
        from json_to_excel import load_items, items_to_excel
    except Exception as e:
        print("Error: can't import json_to_excel helpers:", e)
        return

    try:
        items = load_items(Path(NDJSON_FILE))
    except Exception as e:
        print("Error reading NDJSON for Excel generation:", e)
        return

    if not items:
        # nothing to write
        return

    try:
        items_to_excel(items, Path(OUTPUT_FILE), PROTECT_PASSWORD)
        print(f"Auto Excel: wrote {OUTPUT_FILE}")
    except Exception as e:
        print("Error while generating Excel:", e)


def _auto_excel_worker():
    """Background worker that wakes on events or interval and regenerates Excel."""
    interval = AUTO_EXCEL_INTERVAL
    while not auto_excel_stop.is_set():
        # wait until either event is set (new data) or timeout
        auto_excel_event.wait(interval)
        if auto_excel_stop.is_set():
            break
        # clear the event and generate
        auto_excel_event.clear()
        try:
            generate_excel()
        except Exception as e:
            print("Auto Excel worker error:", e)


def start_auto_excel():
    global auto_excel_thread
    if not AUTO_EXCEL:
        return
    if auto_excel_thread and auto_excel_thread.is_alive():
        return
    auto_excel_stop.clear()
    auto_excel_thread = threading.Thread(target=_auto_excel_worker, daemon=True)
    auto_excel_thread.start()


def stop_auto_excel():
    auto_excel_stop.set()
    auto_excel_event.set()
    if auto_excel_thread:
        auto_excel_thread.join(timeout=5)


def fetch_result(htno):
    payload = {
        "mbstatus": "SEARCH",
        "htno": htno,
        "Submit.x": "25",
        "Submit.y": "8",
    }

    response = session.post(
        URL, data=payload, headers=HEADERS, timeout=15, verify=False
    )

    if "Personal Details" not in response.text:
        return None

    soup = BeautifulSoup(response.text, "html.parser")

    table3 = soup.find(id="AutoNumber3")
    if not table3:
        return None

    rows = table3.find_all("tr")
    try:
        student = {
            "hallticket": rows[1].find_all("td")[1].get_text(strip=True),
            "gender": rows[1].find_all("td")[3].get_text(strip=True),
            "name": rows[2].find_all("td")[1].get_text(strip=True),
            "father": rows[2].find_all("td")[3].get_text(strip=True),
            "course": rows[3].find_all("td")[1].get_text(strip=True),
        }
    except Exception:
        return None

    marks = []
    table4 = soup.find(id="AutoNumber4")
    if table4:
        for row in table4.find_all("tr")[2:]:
            cols = row.find_all("td")
            if len(cols) >= 4:
                marks.append({
                    "code": cols[0].get_text(strip=True),
                    "subject": cols[1].get_text(strip=True),
                    "credits": cols[2].get_text(strip=True),
                    "grade": cols[3].get_text(strip=True),
                })

    final_result = None
    table5 = soup.find(id="AutoNumber5")
    if table5:
        rows5 = table5.find_all("tr")
        if len(rows5) > 2:
            cols = rows5[2].find_all("td")
            if len(cols) > 2:
                final_result = cols[2].get_text(strip=True)

    return {"student": student, "marks": marks, "result": final_result}


def worker(htno):
    print(f"Fetching {htno} ...")
    try:
        data = fetch_result(htno)
        if data:
            # append immediately to disk for visibility and durability
            results.append(data)
            append_result(data)
            print(f"  ✔ SAVED {htno}")
        else:
            # Save a placeholder so missing halltickets are represented in outputs
            placeholder = {
                "student": {"hallticket": htno},
                "marks": [],
                "result": None,
                "status": "NO_RESULT",
            }
            results.append(placeholder)
            append_result(placeholder)
            print(f"  ⚠ NO RESULT {htno} — placeholder saved")
    except Exception as e:
        print(f"  ✖ ERROR {htno}: {e}")


def main():
    # Declare globals early to avoid SyntaxError when module-level names are referenced as defaults
    global URL, START_HT, END_HT, MAX_WORKERS, OUTPUT_FILE, HEADERS, NDJSON_FILE, AUTO_EXCEL, AUTO_EXCEL_INTERVAL, PROTECT_PASSWORD

    parser = argparse.ArgumentParser(description="Fetch exam results by hall ticket range.")
    parser.add_argument("--url", help="Exam result page URL (interactive prompt shown if omitted)", default=None)
    parser.add_argument("--start", type=int, help="Start hall ticket number", default=None)
    parser.add_argument("--end", type=int, help="End hall ticket number", default=None)
    parser.add_argument("--max-workers", type=int, default=MAX_WORKERS, help="Thread pool size")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Output Excel (.xlsx) file")
    parser.add_argument("--no-auto-excel", dest="auto_excel", action="store_false", help="Disable automatic Excel generation during the run")
    parser.add_argument("--auto-excel-interval", type=int, default=AUTO_EXCEL_INTERVAL, help="Auto-Excel worker interval in seconds")
    parser.add_argument("--protect-password", help="Password to protect the grades sheet (optional)", default=None)
    args = parser.parse_args()

    # Set runtime globals from args
    AUTO_EXCEL = args.auto_excel
    AUTO_EXCEL_INTERVAL = args.auto_excel_interval
    PROTECT_PASSWORD = args.protect_password

    # Interactive prompts if arguments are not provided (prompt order: URL -> start -> end)
    def prompt_url(prompt):
        while True:
            val = input(f"{prompt}: ").strip()
            if val == "":
                print("Please enter a URL (required).")
                continue
            if val.startswith("http://") or val.startswith("https://"):
                return val
            print("Please enter a valid URL starting with http:// or https://")

    def prompt_int(prompt, default):
        while True:
            val = input(f"{prompt} [{default}]: ").strip()
            if val == "":
                return default
            try:
                return int(val)
            except ValueError:
                print("Please enter a valid integer.")

    # URL (first) — required when not provided via CLI
    url = args.url if args.url else prompt_url("Enter exam URL")

    # Start and End (prompted in order)
    start = args.start if args.start is not None else prompt_int("Enter start hall ticket number", START_HT)
    end = args.end if args.end is not None else prompt_int("Enter end hall ticket number", END_HT)

    # If start > end, reprompt end until valid
    while start > end:
        print("End hall ticket number must be >= start. Please enter the end hall ticket number again.")
        end = prompt_int("Enter end hall ticket number", END_HT)

    max_workers = args.max_workers
    output_file = args.output

    # Update globals used throughout the program
    URL = url
    START_HT = start
    END_HT = end
    MAX_WORKERS = max_workers

    # Ensure output is Excel (.xlsx) — Excel is mandatory and the default
    if not output_file.lower().endswith('.xlsx'):
        print("Output must be an .xlsx file. Appending .xlsx to the filename.")
        output_file = output_file + '.xlsx'

    OUTPUT_FILE = output_file
    # NDJSON staging file sits alongside the final output
    NDJSON_FILE = os.path.splitext(OUTPUT_FILE)[0] + ".ndjson"
    HEADERS["Referer"] = URL

    print(f"\nUsing URL   : {URL}")
    print(f"Hall tickets: {START_HT} -> {END_HT}")
    print(f"Workers     : {MAX_WORKERS}")
    print(f"Output file : {OUTPUT_FILE}\n")

    print("Clearing old results...")
    clear_old_results()

    # Start the auto-excel thread (if enabled)
    start_auto_excel()

    ht_numbers = [str(ht) for ht in range(START_HT, END_HT + 1)]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(worker, htno) for htno in ht_numbers]
        for _ in as_completed(futures):
            pass

    # Stop the auto-excel worker before final conversion
    stop_auto_excel()

    # Convert NDJSON into a final JSON array for compatibility
    print("\nConverting incremental results to final JSON array...")
    final = []
    with open(NDJSON_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                final.append(json.loads(line))
            except Exception:
                continue

    # Convert final list into richer Excel sheets: flattened results, marks, and raw JSON
    def flatten_top(item):
        flat = {}
        for k, v in item.items():
            if isinstance(v, dict):
                for k2, v2 in v.items():
                    flat[f"{k}.{k2}"] = v2
            elif isinstance(v, list):
                continue
            else:
                flat[k] = v
        return flat

    results_rows = []
    marks_rows = []
    for item in final:
        flat = flatten_top(item)
        # Add explicit top-level hallticket for joining later
        flat["hallticket"] = (item.get("student") or {}).get("hallticket")
        # Do NOT include _raw_json column as requested
        results_rows.append(flat)

        for m in item.get("marks", []):
            mark_row = {"hallticket": (item.get("student") or {}).get("hallticket")}
            for mk, mv in (m or {}).items():
                mark_row[mk] = mv
            marks_rows.append(mark_row)

    df_results = pd.DataFrame(results_rows)
    df_marks = pd.DataFrame(marks_rows)
    raw_df = pd.DataFrame({"json": [json.dumps(i, ensure_ascii=False) for i in final]})

    # Ensure mark sheet has standard columns
    expected_mark_cols = ["hallticket", "code", "subject", "credits", "grade"]
    if df_marks.empty:
        df_marks = pd.DataFrame(columns=expected_mark_cols)
    else:
        df_marks = df_marks.reindex(columns=expected_mark_cols).fillna("")

    # Pivot marks to create per-subject columns for grade, credits, subject
    if not df_marks.empty:
        df_grade_pivot = df_marks.pivot_table(index="hallticket", columns="code", values="grade", aggfunc="first")
        df_credits_pivot = df_marks.pivot_table(index="hallticket", columns="code", values="credits", aggfunc="first")
        df_subject_pivot = df_marks.pivot_table(index="hallticket", columns="code", values="subject", aggfunc="first")

        # Rename columns to include type prefixes
        df_grade_pivot.columns = [f"grade_{str(c)}" for c in df_grade_pivot.columns]
        df_credits_pivot.columns = [f"credits_{str(c)}" for c in df_credits_pivot.columns]
        df_subject_pivot.columns = [f"subject_{str(c)}" for c in df_subject_pivot.columns]

        # Merge pivots into df_results on hallticket
        df_results = df_results.merge(df_grade_pivot.reset_index(), on="hallticket", how="left")
        df_results = df_results.merge(df_credits_pivot.reset_index(), on="hallticket", how="left")
        df_results = df_results.merge(df_subject_pivot.reset_index(), on="hallticket", how="left")

        # Reorder columns: core fields first, then subject columns ordered by frequency (common first, uncommon last)
        core_pref = ["hallticket", "student.name", "student.father", "student.gender", "student.course", "result", "status"]
        core_cols = [c for c in core_pref if c in df_results.columns]
        code_counts = df_marks['code'].value_counts() if not df_marks.empty else []
        codes_sorted = list(code_counts.index) if len(code_counts) else []

        subject_cols = []
        for code in codes_sorted:
            for prefix in (f"grade_{code}", f"credits_{code}", f"subject_{code}"):
                if prefix in df_results.columns:
                    subject_cols.append(prefix)
        remaining = [c for c in df_results.columns if (c.startswith('grade_') or c.startswith('credits_') or c.startswith('subject_')) and c not in subject_cols]
        subject_cols.extend(sorted(remaining))

        other_cols = [c for c in df_results.columns if c not in core_cols + subject_cols]
        final_cols = core_cols + subject_cols + other_cols
        df_results = df_results.reindex(columns=final_cols)

        # Sort by hallticket ascending when possible
        try:
            df_results['hallticket_sort'] = df_results['hallticket'].astype(int)
            df_results = df_results.sort_values('hallticket_sort').drop(columns=['hallticket_sort'])
        except Exception:
            df_results = df_results.sort_values('hallticket')

    # Write sheets in the workbook
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_results.to_excel(writer, sheet_name="results", index=False)
        # also write grades pivot (just grades) for convenience
        if not df_marks.empty:
            df_grade_only = df_grade_pivot.reset_index()
            df_grade_only.to_excel(writer, sheet_name="grades", index=False)
        else:
            pd.DataFrame(columns=["hallticket"]).to_excel(writer, sheet_name="grades", index=False)
        df_marks.to_excel(writer, sheet_name="marks", index=False)
        raw_df.to_excel(writer, sheet_name="raw", index=False)

    # Optionally protect grades sheet (lock grade cells)
    if args.protect_password:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Protection
        except Exception:
            print("Error: openpyxl required for protecting the spreadsheet.")
            return

        wb = load_workbook(OUTPUT_FILE)
        if "grades" in wb.sheetnames:
            ws = wb["grades"]
            # Unlock all cells first
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)
            # Lock grade cells (all except header and hallticket column)
            max_col = ws.max_column
            max_row = ws.max_row
            for r in range(2, max_row + 1):
                for c in range(2, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.protection = Protection(locked=True)
            ws.protection.sheet = True
            ws.protection.set_password(args.protect_password)
            wb.save(OUTPUT_FILE)

    print(f"\nDONE. Excel saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()